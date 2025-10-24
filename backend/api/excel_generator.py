"""
Generador de reportes Excel para CacaoScan.
"""
import logging
import io
from datetime import datetime
from django.utils import timezone
from django.db.models import Count, Avg, Sum
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart

from .models import (
    CacaoImage, CacaoPrediction, Finca, Lote, 
    Notification, ActivityLog, LoginHistory, ReporteGenerado
)

logger = logging.getLogger("cacaoscan.api")


class CacaoReportExcelGenerator:
    """
    Generador de reportes Excel avanzado para CacaoScan.
    """
    
    def __init__(self):
        self.workbook = None
        self.ws = None
    
    def generate_quality_report(self, user, filtros=None):
        """
        Generar reporte de calidad en Excel.
        
        Args:
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar
        """
        try:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Reporte de Calidad"
            
            # Aplicar filtros
            queryset = self._apply_filters(CacaoPrediction.objects.all(), filtros)
            
            # Crear encabezado
            self._create_header("Reporte de Calidad de Granos de Cacao", user)
            
            # Estadísticas generales
            stats = self._get_quality_stats(queryset)
            self._create_stats_section(stats)
            
            # Tabla de análisis detallados
            self._create_detailed_analyses_table(queryset)
            
            # Gráfico de distribución de calidad
            self._create_quality_chart(stats)
            
            # Hoja de resumen
            self._create_summary_sheet(stats, user)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de calidad: {e}")
            raise
    
    def generate_finca_report(self, finca_id, user, filtros=None):
        """
        Generar reporte de finca en Excel.
        
        Args:
            finca_id: ID de la finca
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar
        """
        try:
            finca = Finca.objects.get(id=finca_id)
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = f"Finca {finca.nombre}"
            
            # Crear encabezado
            self._create_header(f"Reporte de Finca: {finca.nombre}", user)
            
            # Información de la finca
            self._create_finca_info_section(finca)
            
            # Estadísticas de lotes
            lotes_stats = self._get_lotes_stats(finca)
            self._create_lotes_stats_section(lotes_stats)
            
            # Análisis por lote
            self._create_lotes_analysis_section(finca)
            
            # Hoja de análisis detallados
            self._create_detailed_lotes_sheet(finca)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de finca: {e}")
            raise
    
    def generate_audit_report(self, user, filtros=None):
        """
        Generar reporte de auditoría en Excel.
        
        Args:
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar
        """
        try:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Auditoría"
            
            # Crear encabezado
            self._create_header("Reporte de Auditoría del Sistema", user)
            
            # Estadísticas de actividad
            activity_stats = self._get_activity_stats(filtros)
            self._create_activity_stats_section(activity_stats)
            
            # Estadísticas de logins
            login_stats = self._get_login_stats(filtros)
            self._create_login_stats_section(login_stats)
            
            # Hoja de actividades detalladas
            self._create_detailed_activities_sheet(filtros)
            
            # Hoja de logins detallados
            self._create_detailed_logins_sheet(filtros)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de auditoría: {e}")
            raise
    
    def generate_custom_report(self, user, tipo_reporte, parametros, filtros=None):
        """
        Generar reporte personalizado en Excel.
        
        Args:
            user: Usuario que solicita el reporte
            tipo_reporte: Tipo de reporte
            parametros: Parámetros del reporte
            filtros: Filtros a aplicar
        """
        try:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Reporte Personalizado"
            
            # Crear encabezado
            self._create_header(f"Reporte Personalizado: {tipo_reporte}", user)
            
            # Generar según tipo
            if tipo_reporte == 'calidad':
                queryset = self._apply_filters(CacaoPrediction.objects.all(), filtros)
                stats = self._get_quality_stats(queryset)
                self._create_custom_quality_section(stats, parametros)
            elif tipo_reporte == 'finca':
                if parametros.get('finca_id'):
                    finca = Finca.objects.get(id=parametros['finca_id'])
                    self._create_custom_finca_section(finca, parametros)
            elif tipo_reporte == 'auditoria':
                activity_stats = self._get_activity_stats(filtros)
                self._create_custom_audit_section(activity_stats, parametros)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel personalizado: {e}")
            raise
    
    def _apply_filters(self, queryset, filtros):
        """Aplicar filtros al queryset."""
        if not filtros:
            return queryset
        
        # Filtro por fecha
        if filtros.get('fecha_desde'):
            queryset = queryset.filter(created_at__date__gte=filtros['fecha_desde'])
        if filtros.get('fecha_hasta'):
            queryset = queryset.filter(created_at__date__lte=filtros['fecha_hasta'])
        
        # Filtro por usuario
        if filtros.get('usuario_id'):
            queryset = queryset.filter(image__user_id=filtros['usuario_id'])
        
        # Filtro por finca
        if filtros.get('finca_id'):
            queryset = queryset.filter(image__finca=filtros['finca_id'])
        
        # Filtro por lote
        if filtros.get('lote_id'):
            queryset = queryset.filter(image__lote_id=filtros['lote_id'])
        
        return queryset
    
    def _create_header(self, title, user):
        """Crear encabezado del reporte."""
        # Título principal
        self.ws['A1'] = title
        self.ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        self.ws['A1'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('A1:F1')
        
        # Información del reporte
        self.ws['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        self.ws['A3'].font = Font(size=10, italic=True)
        
        self.ws['A4'] = f"Usuario: {user.get_full_name() or user.username}"
        self.ws['A4'].font = Font(size=10, italic=True)
        
        # Espacio
        self.ws['A6'] = ""
    
    def _get_quality_stats(self, queryset):
        """Obtener estadísticas de calidad."""
        total_analyses = queryset.count()
        
        if total_analyses == 0:
            return {
                'total_analyses': 0,
                'avg_confidence': 0,
                'quality_distribution': {},
                'avg_dimensions': {},
                'avg_weight': 0
            }
        
        # Estadísticas de confianza
        avg_confidence = queryset.aggregate(avg=Avg('average_confidence'))['avg'] or 0
        
        # Distribución de calidad
        quality_distribution = {
            'Excelente (≥90%)': queryset.filter(average_confidence__gte=0.9).count(),
            'Buena (80-89%)': queryset.filter(average_confidence__gte=0.8, average_confidence__lt=0.9).count(),
            'Regular (70-79%)': queryset.filter(average_confidence__gte=0.7, average_confidence__lt=0.8).count(),
            'Baja (<70%)': queryset.filter(average_confidence__lt=0.7).count(),
        }
        
        # Dimensiones promedio
        avg_dimensions = queryset.aggregate(
            avg_alto=Avg('alto_mm'),
            avg_ancho=Avg('ancho_mm'),
            avg_grosor=Avg('grosor_mm')
        )
        
        # Peso promedio
        avg_weight = queryset.aggregate(avg=Avg('peso_g'))['avg'] or 0
        
        return {
            'total_analyses': total_analyses,
            'avg_confidence': round(float(avg_confidence) * 100, 2),
            'quality_distribution': quality_distribution,
            'avg_dimensions': {
                'alto': round(float(avg_dimensions['avg_alto'] or 0), 2),
                'ancho': round(float(avg_dimensions['avg_ancho'] or 0), 2),
                'grosor': round(float(avg_dimensions['avg_grosor'] or 0), 2),
            },
            'avg_weight': round(float(avg_weight), 2)
        }
    
    def _create_stats_section(self, stats):
        """Crear sección de estadísticas."""
        # Título de sección
        self.ws['A8'] = "Estadísticas Generales"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Datos de estadísticas
        data = [
            ['Métrica', 'Valor'],
            ['Total de Análisis', stats['total_analyses']],
            ['Confianza Promedio', f"{stats['avg_confidence']}%"],
            ['Alto Promedio', f"{stats['avg_dimensions']['alto']} mm"],
            ['Ancho Promedio', f"{stats['avg_dimensions']['ancho']} mm"],
            ['Grosor Promedio', f"{stats['avg_dimensions']['grosor']} mm"],
            ['Peso Promedio', f"{stats['avg_weight']} g"],
        ]
        
        # Crear tabla
        for row_num, row_data in enumerate(data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
